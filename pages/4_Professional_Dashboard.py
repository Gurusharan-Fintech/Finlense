import streamlit as st
import pandas as pd
import yfinance as yf
import plotly.graph_objects as go
import io
import requests
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, StockChart

# -------------------------------
# AI Analysis (Ollama Fallback)
# -------------------------------
def generate_ai_analysis(ticker, df):
    try:
        import ollama
        latest_close = df["Close"].iloc[-1] if not df.empty else "N/A"
        prompt = f"""
        Provide a detailed financial analysis for stock {ticker}.
        Latest Close Price: {latest_close}
        Include trends, risks, opportunities, and investor outlook.
        """
        response = ollama.chat(model="mistral", messages=[{"role": "user", "content": prompt}])
        return response["message"]["content"].strip()
    except Exception:
        return "⚠️ AI analysis unavailable (Ollama not installed). Please install it from https://ollama.ai/."

# -------------------------------
# Create Excel with Charts
# -------------------------------
def create_excel(ticker, df, ai_text):
    buffer = io.BytesIO()
    df_reset = df.reset_index()

    # Step 1: Write to Excel
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_reset.to_excel(writer, sheet_name="Stock Data", index=False)

        summary_df = pd.DataFrame({
            "Metric": ["Ticker", "Rows", "Columns"],
            "Value": [ticker, df_reset.shape[0], df_reset.shape[1]]
        })
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ai_df = pd.DataFrame({"AI Insights": [ai_text]})
        ai_df.to_excel(writer, sheet_name="AI Analysis", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)

    # Step 2: Auto-adjust columns
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    # Step 3: Add Charts
    if not df.empty:
        ws_data = wb["Stock Data"]
        ws_chart = wb.create_sheet("Charts")

        # Line Chart: Closing Price + MA20
        line_chart = LineChart()
        line_chart.title = f"{ticker} Closing Price with 20-Day MA"
        values = Reference(ws_data, min_col=df_reset.columns.get_loc("Close")+1,
                           min_row=2, max_row=len(df_reset)+1)
        dates = Reference(ws_data, min_col=1, min_row=2, max_row=len(df_reset)+1)
        line_chart.add_data(values, titles_from_data=False)
        line_chart.set_categories(dates)

        # Add MA20
        df_reset["MA20"] = df_reset["Close"].rolling(20).mean()
        ma_col = df_reset.shape[1] + 1
        ws_data.cell(row=1, column=ma_col, value="MA20")
        for i, val in enumerate(df_reset["MA20"], start=2):
            ws_data.cell(row=i, column=ma_col, value=float(val) if pd.notna(val) else None)
        ma_values = Reference(ws_data, min_col=ma_col, min_row=2, max_row=len(df_reset)+1)
        line_chart.add_data(ma_values, titles_from_data=False)

        ws_chart.add_chart(line_chart, "B2")

        # Stock Chart (OHLC)
        stock_chart = StockChart()
        stock_chart.title = f"{ticker} OHLC"
        ohlc = Reference(ws_data, min_col=df_reset.columns.get_loc("Open")+1, max_col=df_reset.columns.get_loc("Close")+1,
                         min_row=1, max_row=len(df_reset)+1)
        stock_chart.add_data(ohlc, titles_from_data=True)
        stock_chart.set_categories(dates)
        ws_chart.add_chart(stock_chart, "B20")

    buffer2 = io.BytesIO()
    wb.save(buffer2)
    buffer2.seek(0)
    return buffer2

# -------------------------------
# Create PDF Report
# -------------------------------
def create_pdf(ticker, df, ai_text):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph(f"<b>Stock Report: {ticker}</b>", styles["Title"]))
    content.append(Spacer(1, 12))

    content.append(Paragraph("<b>AI Stock Analysis</b>", styles["Heading2"]))
    content.append(Paragraph(ai_text, styles["Normal"]))
    content.append(Spacer(1, 12))

    if not df.empty:
        stats = [
            ["Latest Close", f"{df['Close'].iloc[-1]:.2f}"],
            ["Max Close", f"{df['Close'].max():.2f}"],
            ["Min Close", f"{df['Close'].min():.2f}"],
        ]
        table = Table(stats)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                   ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                   ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                   ("GRID", (0, 0), (-1, -1), 1, colors.black)]))
        content.append(table)

    doc.build(content)
    buffer.seek(0)
    return buffer

# -------------------------------
# Main Dashboard
# -------------------------------
def main():
    st.title("📊 Professional Data & Trends")

    ticker = st.text_input("Enter Stock Ticker (e.g. AAPL, TSLA):", "AAPL")
    if not ticker:
        return

    df = yf.download(ticker, period="6mo", interval="1d")

    if df.empty:
        st.warning("⚠️ No data found for this ticker.")
        return

    # AI Analysis
    st.subheader("😁 AI Stock Analysis")
    ai_text = generate_ai_analysis(ticker, df)
    st.write(ai_text)

    # Price Movement Chart
    st.subheader("📈 Price Movement")
    fig = go.Figure()
    fig.add_trace(go.Candlestick(
        x=df.index,
        open=df["Open"], high=df["High"],
        low=df["Low"], close=df["Close"],
        name="OHLC"
    ))
    fig.add_trace(go.Scatter(x=df.index, y=df["Close"].rolling(20).mean(),
                             line=dict(color="blue", width=1.5), name="20D MA"))
    st.plotly_chart(fig, use_container_width=True)

    # Volume Chart
    st.subheader("📊 Volume Trends")
    vol_fig = go.Figure()
    vol_fig.add_trace(go.Bar(x=df.index, y=df["Volume"], name="Volume"))
    st.plotly_chart(vol_fig, use_container_width=True)

    # Downloads
    pdf_buffer = create_pdf(ticker, df, ai_text)
    excel_buffer = create_excel(ticker, df, ai_text)

    st.download_button("📥 Download PDF Report", data=pdf_buffer,
                       file_name=f"{ticker}_report.pdf", mime="application/pdf")

    st.download_button("📥 Download Excel Report", data=excel_buffer,
                       file_name=f"{ticker}_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    main()
